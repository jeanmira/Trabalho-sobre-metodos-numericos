# ------------------------------- /usr/bin/g++-7 ------------------------------#
# ------------------------------- coding: utf-8 -------------------------------#
# Criado por:   Jean Marcelo Mira Junior
#               Lucas Daniel dos Santos
# Versão: 1.0
# Criado em: 13/04/2021
# Sistema operacional: Linux - Ubuntu 20.04.1 LTS
# Python 3
# ------------------------------ Pacotes --------------------------------------#
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
# -----------------------------------------------------------------------------#


def edoEuler(f, r0, s0, nPassos, h):
    """
    Método numérico de Euler que consiste em
    utilizar as retas tangentes para encontrar
    o valor da função em um ponto próximo.

    Recebe:
    f       => Função que vai ser resolvida
    r0      => Armazena o valor da função inicial
    s0      => Armazena os passos de tempo inicial
    nPassos => Quantidade de passos de tempo
    h       => Distância

    Retorna:
    s       => Variável  independente
    r       => Variável/Função  dependente
    """

    nEquacoes = len(r0)

    r = np.zeros([nPassos, nEquacoes], dtype=np.float32)
    s = np.zeros(nPassos, dtype=np.float32)

    r[0] = r0
    s[0] = s0

    for n in range(0, nPassos - 1):
        k1 = f(s[n], r[n])
        r[n + 1] = r[n] + k1 * h
        s[n + 1] = s[n] + h

    return(s, r)


def edoHeun(f, r0, s0, nPassos, h):
    """
    Método numérico de Heun é o método
    numérico de Euler melhorado.

    Recebe:
    f       => Função que vai ser resolvida
    y0      => Armazena o valor da função inicial
    s0      => Armazena os passos de tempo inicial
    nPassos => Quantidade de passos de tempo
    h       => Distância

    Retorna:
    s       => Variável  independente
    r       => Variável/Função  dependente
    """

    nEquacoes = len(r0)

    r = np.zeros([nPassos, nEquacoes], dtype=np.float32)
    s = np.zeros(nPassos, dtype=np.float32)

    r[0] = r0
    s[0] = s0

    for n in range(0, nPassos - 1):
        s[n + 1] = s[n] + h
        k1 = f(s[n], r[n])
        k2 = f(s[n + 1], r[n] + h * k1)
        r[n + 1] = r[n] + (1 / 2) * (k1 + k2) * h

    return(s, r)


def edoRungeKutta(f, r0, s0, nPassos, h):
    """
    Método numérico de Runge-Kutta é obtido
    utilizando uma expansão em série de Taylor
    solicitando um erro local de determinada ordem.

    Recebe:
    f       => Função que vai ser resolvida
    y0      => Armazena o valor da função inicial
    s0      => Armazena os passos de tempo inicial
    nPassos => Quantidade de passos de tempo
    h       => Distância

    Retorna:
    s       => Variável  independente
    r       => Variável/Função  dependente
    """

    nEquacoes = len(r0)
    r = np.zeros([nPassos, nEquacoes], dtype=np.float32)
    s = np.zeros(nPassos, dtype=np.float32)

    r[0] = r0
    s[0] = s0

    for n in range(0, nPassos - 1):
        s[n + 1] = s[n] + h
        k1 = f(s[n], r[n])
        k2 = f(s[n] + h * (1 / 2), r[n] + h * k1 * (1 / 2))
        k3 = f(s[n] + h * (1 / 2), r[n] + h * k2 * (1 / 2))
        k4 = f(s[n + 1], r[n] + h * k3)
        r[n + 1] = r[n] + h*(k1 + 2 * k2 + 2 * k3 + k4) * (1 / 6)

    return(s, r)


def edoRungeKuttaFehlberg(f, r0, s0, nPassos, h, tol, k, alpha):
    """
    Método numérico de Runge-Kutta-Fehlberg é obtido
    utilizando uma expansão em série de Taylor
    solicitando um erro local de determinada ordem,
    com o passo de h variável.

    Recebe:
    f       => Função que vai ser resolvida
    y0      => Armazena o valor da função inicial
    s0      => Armazena os passos de tempo inicial
    nPassos => Quantidade de passos de tempo
    h       => Distância variante
    tol     => Tolerancia de erro
    k       => Ordem do método
    a       => Valor alpha

    Retorna:
    s       => Variável  independente
    r       => Variável/Função  dependente
    """

    nEquacoes = len(r0)

    r = np.zeros([nPassos, nEquacoes], dtype=np.float32)
    s = np.zeros(nPassos, dtype=np.float32)
    q = np.zeros(nPassos, dtype=np.float32)
    y4 = np.zeros(nPassos, dtype=np.float32)
    y5 = np.zeros(nPassos, dtype=np.float32)

    r[0] = r0
    s[0] = s0

    qMin = 0

    for n in range(0, nPassos - 1):
        while (True):
            k1 = f(s[n], r[n])
            k2 = f(s[n] + h * (1 / 4), r[n] + h * k1 * (1 / 4))
            k3 = f(s[n] + h * (3 / 8), r[n] + h * ((k1 * 3 + k2 * 9) / 32))
            k4 = f(s[n] + h * (12 / 13), r[n] + h *
                   ((k1 * 1932 - k2 * 7200 + k3 * 7296) / 2197))
            k5 = f(s[n] + h, r[n] + h * (k1 * (439 / 216) - k2 *
                                         8 + k3 * (3680 / 513) - k4 * (845 / 4104)))
            k6 = f(s[n] + h * (1 / 2), r[n] + h * (- k1 * (8 / 27) + k2 *
                                                   2 - k3 * (3544 / 2565) + k4 * (1859 / 4104) - k5 * (11 / 40)))
            y4 = (r[n] + h * (k1 * (25 / 216) + k3 *
                              (1408 / 2565) + k4 * (2197 / 4104) - k5 * (1 / 5)))
            y5 = (r[n] + h * (k1 * (16 / 135) + k3 * (6656 / 12825) +
                              k4 * (28561 / 56430) - k5 * (9 / 50) + k6 * (2 / 55)))
            q = alpha * ((tol * h) / abs(y4 - y5)) ** (1 / k)

            qMin = min(q)

            if(qMin >= 1):
                break
            else:
                h = qMin * h

        s[n + 1] = s[n] + h
        r[n + 1] = y5
        h = qMin * h

    return(s, r)


def grafico(se, re, sh, rh, sr, rr, sf, rf):
    """
    Salva os dados em uma planilha

    Recebe:
    se      =>  Variável  independente
    re      =>  Dados de Euler
    sh      =>  Variável  independente
    rh      =>  Dados de Heun
    sr      =>  Variável  independente
    rr      =>  Dados de Runge-Kutta
    sf      =>  Variável  independente
    rf      =>  Dados de Runge-Kutta-Fehlberg

    Retorna: NADA
    """

    fi1, r1, z1 = re.T
    fi2, r2, z2 = rh.T
    fi3, r3, z3 = rr.T
    fi4, r4, z4 = rf.T

    fig, (fi, r, z, gota) = plt.subplots(
        nrows=1, ncols=4, sharex=False, sharey=False, figsize=(16, 10))

    fi.set_title('Gráfico de fi x s')
    fi.set_xlabel("s")
    fi.set_ylabel("fi")
    fi.plot(se, fi1,  'g.', color='red', label='Euler')
    fi.plot(sh, fi2,  'g.', color='green', label='Heun')
    fi.plot(sr, fi3, 'g.', color='blue', label='Runge-Kutta')
    fi.plot(sf, fi4,  'g.', color='yellow', label='Runge-Kutta-Fehlberg')
    # fi.legend()

    r.set_title('Gráfico de r x s')
    r.set_xlabel("s")
    r.set_ylabel("r")
    r.plot(se, r1,  'g.', color='red', label='Euler')
    r.plot(sh, r2,  'g.', color='green', label='Heun')
    r.plot(sr, r3,  'g.', color='blue', label='Runge-Kutta')
    r.plot(sf, r4,  'g.', color='yellow', label='Runge-Kutta-Fehlberg')
    # r.legend()

    z.set_title('Gráfico de z x s')
    z.set_xlabel("s")
    z.set_ylabel("z")
    z.plot(se, z1,  'g.', color='red', label='Euler')
    z.plot(sh, z2,  'g.', color='green', label='Heun')
    z.plot(sr, z3,  'g.', color='blue', label='Runge-Kutta')
    z.plot(sf, z4,  'g.', color='yellow', label='Runge-Kutta-Fehlberg')
    # z.legend()

    gota.set_title("Gráfico da Gota - z x r")
    gota.set_aspect(1)
    gota.set_xlabel("r")
    gota.set_ylabel("z")
    gota.plot(r1, z1, 'g.', color='red', label='Euler')
    gota.plot(r2, z2, 'g.', color='green', label='Heun')
    gota.plot(r3, z3, 'g.', color='blue', label='Runge-Kutta')
    gota.plot(r4, z4, 'g.', color='yellow', label='Runge-Kutta-Fehlberg')
    gota.legend(loc='upper right', bbox_to_anchor=(1, -0.3))

    plt.subplot_tool()
    fig.suptitle('Resultados')
    # plt.saveaxs("resultados.png")
    plt.show()


def gota(re, rh, rr, rf):
    """
    Faz o grafico da gota (z x r)

    Recebe:
    re      =>  Dados de Euler
    rh      =>  Dados de Heun
    rr      =>  Dados de Runge-Kutta
    rf      =>  Dados de Runge-Kutta-Fehlberg

    Retorna: NADA
    """

    fi1, r1, z1 = re.T
    fi2, r2, z2 = rh.T
    fi3, r3, z3 = rr.T
    fi4, r4, z4 = rf.T

    gota = plt.subplot(1, 1, 1)
    gota.set_title("Gráfico da Gota - z x r")
    gota.set_aspect(1)
    gota.plot(r1, z1, 'g--', color='red', label='Euler')
    gota.plot(r2, z2, 'g--', color='green', label='Heun')
    gota.plot(r3, z3, 'g--', color='blue', label='Runge-Kutta')
    gota.plot(r4, z4, 'g--', color='yellow', label='Runge-Kutta-Fehlberg')
    gota.legend(loc='upper right', bbox_to_anchor=(-0.3, 1.0))
    # plt.saveaxs("r_z.png")
    plt.show()


def planilha(nPassos, se, re, sh, rh, sr, rr, sf, rf):
    """
    Salva os dados em uma planilha

    Recebe:
    se      =>  Variável  independente
    re      =>  Dados de Euler
    sh      =>  Variável  independente
    rh      =>  Dados de Heun
    sr      =>  Variável  independente
    rr      =>  Dados de Runge-Kutta
    sf      =>  Variável  independente
    rf      =>  Dados de Runge-Kutta-Fehlberg

    Retorna: NADA
    """

    fi1, r1, z1 = re.T
    fi2, r2, z2 = rh.T
    fi3, r3, z3 = rr.T
    fi4, r4, z4 = rf.T

    arquivo_excel = Workbook()
    planilha1 = arquivo_excel.active
    planilha1.title = "Dados"
    planilha1['A1'] = 'sE'
    planilha1['B1'] = 'fiE'
    planilha1['C1'] = 'rE'
    planilha1['D1'] = 'zE'
    planilha1['E1'] = 'sH'
    planilha1['F1'] = 'fiH'
    planilha1['G1'] = 'rH'
    planilha1['H1'] = 'zH'
    planilha1['I1'] = 'sR'
    planilha1['J1'] = 'fiR'
    planilha1['K1'] = 'rR'
    planilha1['L1'] = 'zR'
    planilha1['M1'] = 'sF'
    planilha1['N1'] = 'fiF'
    planilha1['O1'] = 'rF'
    planilha1['P1'] = 'zF'

    for i in range(2, nPassos+2):
        planilha1.cell(row=i, column=1, value=float(se[i-2]))
        planilha1.cell(row=i, column=2, value=float(fi1[i-2]))
        planilha1.cell(row=i, column=3, value=float(r1[i-2]))
        planilha1.cell(row=i, column=4, value=float(z1[i-2]))
        # ----------------------------------------------------
        planilha1.cell(row=i, column=5, value=float(sh[i-2]))
        planilha1.cell(row=i, column=6, value=float(fi2[i-2]))
        planilha1.cell(row=i, column=7, value=float(r2[i-2]))
        planilha1.cell(row=i, column=8, value=float(z2[i-2]))
        # ----------------------------------------------------
        planilha1.cell(row=i, column=9, value=float(sr[i-2]))
        planilha1.cell(row=i, column=10, value=float(fi3[i-2]))
        planilha1.cell(row=i, column=11, value=float(r3[i-2]))
        planilha1.cell(row=i, column=12, value=float(z3[i-2]))
        # ----------------------------------------------------
        planilha1.cell(row=i, column=13, value=float(sf[i-2]))
        planilha1.cell(row=i, column=14, value=float(fi4[i-2]))
        planilha1.cell(row=i, column=15, value=float(r4[i-2]))
        planilha1.cell(row=i, column=16, value=float(z4[i-2]))

    arquivo_excel.save("Dados.xlsx")
